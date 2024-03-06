'''
    Transforms the input DataFrame according to specific rules.

    This function finds specific rows and columns in the input DataFrame,
    extracts and renames certain columns, and adds a new column. The function
    is specifically designed to work with data from xtb.

    :param df: The input DataFrame to be transformed.
    :type df: pandas.DataFrame
    :param sheet_name: The name of the sheet from which the DataFrame was extracted.
    :type sheet_name: str
    :return: The transformed DataFrame.
    :rtype: pandas.DataFrame

    :Example:

    >>> df_cash_etl(df, 'Sheet1')
'''
import os
import yaml
import pandas as pd
from openpyxl import load_workbook

def df_cash_etl(df, sheet_name):
    # Find the index of the first row containing 'ID'
    id_index = df[df.isin(['ID']).any(axis=1)].index[0]

    # Set the row at 'ID' index as the header and Select all rows from 'ID' to the end 
    df.columns = df.iloc[id_index]
    df = df.loc[id_index + 1:]

    # Select only the columns with the relevant data
    df = df.iloc[:, [2, 3, 4, 5, 6]]

    # Rename columns by index
    df.columns.values[0] = 'type'
    df.columns.values[1] = 'date'
    df.columns.values[2] = 'comments'
    df.columns.values[3] = 'symbol'
    df.columns.values[4] = 'value'
    df['source'] = 'xtb_' + sheet_name.lower().replace(' ', '_')
    df['currency'] = df['comments'].str.split().str[1]
    df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d').dt.strftime('%d/%m/%Y')
    df['merge_key'] = df['symbol'].str.split('.').str[0] + "|" + df['date'].astype(str)

    # Split the dataframe into deposit and divident
    df_deposit = df.loc[df['type'] == 'Wplata']
    df_deposit['category'] = 'deposit'
    df_deposit.rename(columns={'value': 'deposit'}, inplace=True)
    columns_to_drop = ['type', 'comments', 'symbol', 'date', 'source', 'category', 'currency']
    df_deposit = df_deposit.drop(columns=columns_to_drop)

    df_dividend = df.loc[df['type'] == 'Dywidenda']
    df_dividend['category'] = 'dividend'
    df_dividend.rename(columns={'value': 'dividend'}, inplace=True)
    columns_to_drop = ['type', 'comments', 'symbol', 'date', 'source', 'category']
    df_dividend = df_dividend.drop(columns=columns_to_drop)


    # Create a dictionary to store the dataframes
    result = {'deposit': df_deposit, 'dividend': df_dividend}

    return result

def df_open_etl(df, sheet_name):
    # Find the index of the first row containing 'ID'
    id_index = df[df.isin(['Pozycja']).any(axis=1)].index[0]

    # Set the row at 'ID' index as the header and Select all rows from 'ID' to the end 
    df.columns = df.iloc[id_index]
    df = df.loc[id_index + 1:]

    # Select only the columns with the relevant data
    df = df.iloc[:, [2, 3, 4, 5, 8]]

    # Rename columns by index
    df.columns.values[0] = 'symbol'
    df.columns.values[1] = 'type'
    df.columns.values[2] = 'shares'
    df.columns.values[3] = 'date'
    df.columns.values[4] = 'buy_value'
    df['source'] = 'xtb_' + sheet_name.lower().replace(' ', '_')
    df['category'] = 'open_position'
    df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d').dt.strftime('%d/%m/%Y')
    df['merge_key'] = df['symbol'].str.split('.').str[0] + "|" + df['date'].astype(str)

    # Split the dataframe into deposit and divident
    df = df.loc[df['type'].notnull()]

    # Drop columns
    # columns_to_drop = ['type', 'symbol', 'date', 'source', 'category']
    columns_to_drop = ['type', 'source', 'category']
    df = df.drop(columns=columns_to_drop)

    # Rename columns
    df.rename(columns={'shares': 'open_shares'}, inplace=True)

    # Aggregate both, open_shares and buy_value
    df = df.groupby(['merge_key', 'symbol', 'date']).agg({'open_shares': 'sum', 'buy_value': 'sum'}).reset_index()

    #df = df.groupby(['merge_key', 'symbol', 'date']).agg({'open_shares': 'sum'}).reset_index()

    # Create a dictionary to store the dataframes
    result = {'open': df}

    return result

def df_closed_etl(df, sheet_name):
    # Find the index of the first row containing 'ID'
    id_index = df[df.isin(['Pozycja']).any(axis=1)].index[0]

    # Set the row at 'ID' index as the header and Select all rows from 'ID' to the end 
    df.columns = df.iloc[id_index]
    df = df.loc[id_index + 1:]

    # Select only the columns with the relevant data
    df_open = df.iloc[:, [2, 3, 4, 5, 11]]
    df_closed = df.iloc[:, [2, 3, 4, 7, 12]]

    # Filter the dataframe
    df_open = df_open.loc[df_open['Symbol'].notnull()]
    df_closed = df_closed.loc[df_closed['Symbol'].notnull()]

    # Rename columns by index
    df_open.columns.values[0] = 'symbol'
    df_open.columns.values[1] = 'type'
    df_open.columns.values[2] = 'shares'
    df_open.columns.values[3] = 'date'
    df_open.columns.values[4] = 'buy_value'
    df_open['source'] = 'xtb_' + sheet_name.lower().replace(' ', '_')
    df_open['category'] = 'open_position'
    df_open['date'] = pd.to_datetime(df_open['date'], format='%Y-%m-%d').dt.strftime('%d/%m/%Y')
    df_open['merge_key'] = df_open['symbol'].str.split('.').str[0] + "|" + df_open['date'].astype(str)

    # Drop columns
    #columns_to_drop = ['type', 'symbol', 'date', 'source', 'category']
    columns_to_drop = ['type', 'source', 'category']
    df_open = df_open.drop(columns=columns_to_drop)
    
    # Rename columns
    df_open.rename(columns={'shares': 'open_shares'}, inplace=True)

    # Aggregate both open_shares and buy_value
    df_open = df_open.groupby(['merge_key', 'symbol', 'date']).agg({'open_shares': 'sum', 'buy_value': 'sum'}).reset_index()

    df_closed.columns.values[0] = 'symbol'
    df_closed.columns.values[1] = 'type'
    df_closed.columns.values[2] = 'shares'
    df_closed.columns.values[3] = 'date'
    df_closed.columns.values[4] = 'sell_value'
    df_closed['source'] = 'xtb_' + sheet_name.lower().replace(' ', '_')
    df_closed['category'] = 'closed_position'
    df_closed['date'] = pd.to_datetime(df_closed['date'], format='%Y-%m-%d').dt.strftime('%d/%m/%Y')
    df_closed['merge_key'] = df_closed['symbol'].str.split('.').str[0] + "|" + df_closed['date'].astype(str)

    # Drop columns
    columns_to_drop = ['type', 'source', 'category']
    df_closed = df_closed.drop(columns=columns_to_drop)

    # Rename columns
    df_closed.rename(columns={'shares': 'closed_shares'}, inplace=True)

    # Aggregate both, open_shares and buy_value
    df_closed = df_closed.groupby(['merge_key', 'symbol', 'date']).agg({'closed_shares': 'sum', 'sell_value': 'sum'}).reset_index()

    # Create a dictionary to store the dataframes
    result = {'open': df_open, 'closed': df_closed}

    return result

def get_xtb_main_data():
    """
    Extracts and transforms data from an Excel file.

    This function reads data from the 'CLOSED POSITION HISTORY', 'OPEN POSITION', and 'CASH OPERATION' sheets of the Excel file, transforms the data, and returns it in a dictionary.

    :param xlsx_file: The path to the Excel file.
    :type xlsx_file: str
    :return: A dictionary containing the transformed data from the Excel file.
    :rtype: dict
    """
    # Load the configuration file
    with open('finance_portfel/variables.yaml', 'r') as file:
        config = yaml.safe_load(file)

    # Get the path of the folder containing the CSV file
    folder_path = config['data_source_path']['xtb_path']

    # Check if the folder exists
    if os.path.exists(folder_path):

        # Get a list of all the files in the folder
        files = os.listdir(folder_path)

        # Get the first file in the list
        xlsx_file = os.path.join(folder_path, files[0])

        # Load the workbook and get all sheet names
        wb = load_workbook(filename=xlsx_file, read_only=True)
        sheet_names = wb.sheetnames

        # Find the sheet name that contains the specific text
        sn_closed_position = next((sheet for sheet in sheet_names if sheet == 'CLOSED POSITION HISTORY'), None)
        sn_open_position = next((sheet for sheet in sheet_names if 'OPEN POSITION' in sheet), None)
        sn_cash_operation = next((sheet for sheet in sheet_names if 'CASH OPERATION' in sheet), None)

        # Check if the required sheets are found
        if not sn_closed_position or not sn_open_position or not sn_cash_operation:
            raise ValueError('Required sheets not found in the workbook.')

        # Load the data from the sheets
        df_closed = pd.read_excel(xlsx_file, sheet_name=sn_closed_position)
        df_open = pd.read_excel(xlsx_file, sheet_name=sn_open_position)
        df_cash = pd.read_excel(xlsx_file, sheet_name=sn_cash_operation)
        
        # ETL closed position
        df_closed_open = df_closed_etl(df_closed, sn_closed_position)
        df_closed = df_closed_open['closed']
        df_closed.columns = map(str.lower, df_closed.columns)

        # Change the data type
        df_closed['closed_shares'] = df_closed['closed_shares'].astype('int64')
        df_closed['sell_value'] = df_closed['sell_value'].astype('float64')
        df_closed['date'] = pd.to_datetime(df_closed['date'], format='%d/%m/%Y')
        
        # Sort the dataframe
        df_closed = df_closed.sort_values(['symbol', 'date'])

        # Create a new column 'open_shares_cum'
        df_closed['closed_shares_cum'] = df_closed.groupby('symbol')['closed_shares'].cumsum()
        df_closed['sell_value_cum'] = df_closed.groupby('symbol')['sell_value'].cumsum()

        # Drop columns
        columns_to_drop = ['symbol', 'date']
        df_closed = df_closed.drop(columns=columns_to_drop)

        # ETL open position
        df_open_1 = df_closed_open['open']
        df_open_2 = df_open_etl(df_open, sn_open_position)['open']
        df_open = pd.concat([df_open_1, df_open_2])
        df_open.columns = map(str.lower, df_open.columns)

        # Change the data type
        df_open['open_shares'] = df_open['open_shares'].astype('int64')
        df_open['buy_value'] = df_open['buy_value'].astype('float64')
        df_open['date'] = pd.to_datetime(df_open['date'], format='%d/%m/%Y')
        
        # Sort the dataframe
        df_open = df_open.sort_values(['symbol', 'date'])

        # Create a new column 'open_shares_cum'
        df_open['open_shares_cum'] = df_open.groupby('symbol')['open_shares'].cumsum()
        df_open['buy_value_cum'] = df_open.groupby('symbol')['buy_value'].cumsum()

        # Drop columns
        columns_to_drop = ['symbol', 'date']
        df_open = df_open.drop(columns=columns_to_drop)

        df_deposit = df_cash_etl(df_cash, sn_cash_operation)['deposit']
        df_deposit.columns = map(str.lower, df_deposit.columns)
        df_dividend = df_cash_etl(df_cash, sn_cash_operation)['dividend']
        df_dividend.columns = map(str.lower, df_dividend.columns)
        
        # Create a dictionary to store the dataframes
        result = {'open': df_open, 'closed': df_closed, 'deposit': df_deposit, 'dividend': df_dividend}

    return result

if __name__ == '__main__':
    get_xtb_main_data()

    